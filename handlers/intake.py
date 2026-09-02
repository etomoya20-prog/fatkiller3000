"""Приём ежедневных отчётов о еде: текст, выгрузки из приложений, скриншоты."""

from __future__ import annotations

import datetime as dt
import io
import logging

from aiogram import Bot, F, Router
from aiogram.enums import ChatType
from aiogram.filters import Command
from aiogram.types import Message

import db
import llm
from config import MSK, Config

log = logging.getLogger(__name__)

router = Router(name="intake")
router.message.filter(F.chat.type == ChatType.PRIVATE)

TEXT_EXTENSIONS = {".txt", ".csv", ".tsv", ".json", ".md", ".log"}
MAX_FILE_BYTES = 15 * 1024 * 1024


# Трекеры, которые стоит советовать: с русской базой продуктов и сканером
# штрихкодов, то есть пригодные для ежедневного учёта без мучений.
TRACKERS = "FatSecret, MyFitnessPal или YAZIO"


def today_msk() -> dt.date:
    return dt.datetime.now(MSK).date()


def _photo_estimate_warning(guide_url: str) -> str:
    """Приписка к оценке по фотографии.

    Оценку мы всё равно записываем — пропущенный день хуже неточного, — но
    человек должен понимать, что это черновик, иначе он решит, что фотографии
    достаточно, и будет месяц вести дневник с погрешностью в треть.
    """
    lines = [
        "⚠️ <b>Это прикидка по фотографии.</b> По снимку не видно ни масла на "
        "сковороде, ни точного веса порции — ошибка легко доходит до трети.",
        f"Для настоящего учёта заведи трекер калорий: {TRACKERS}. Там база "
        f"продуктов и сканер штрихкодов. Пришли мне скриншот или выгрузку "
        f"оттуда — запишу точные цифры вместо оценки (/reset сотрёт сегодняшнее).",
    ]
    if guide_url:
        lines.append(f'📖 <a href="{guide_url}">Как вести дневник питания</a>')
    return "\n\n".join(lines)


async def _require_profile(message: Message) -> dict | None:
    profile = await db.get_user(message.from_user.id)
    if not profile or not profile["onboarded_at"]:
        await message.answer(
            "Сначала нужно заполнить анкету — без неё я не знаю твою норму. Жми /start"
        )
        return None
    return dict(profile)


def _extract_text(filename: str, raw: bytes) -> str | None:
    """Достаёт текст из выгрузки. Возвращает None для форматов, которые не понимаем."""
    lowered = filename.lower()

    if any(lowered.endswith(ext) for ext in TEXT_EXTENSIONS):
        for encoding in ("utf-8", "utf-8-sig", "cp1251"):
            try:
                return raw.decode(encoding)
            except UnicodeDecodeError:
                continue
        return raw.decode("utf-8", errors="replace")

    if lowered.endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in workbook.worksheets:
            lines.append(f"# Лист: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                cells = [str(c) for c in row if c is not None]
                if cells:
                    lines.append("\t".join(cells))
        workbook.close()
        return "\n".join(lines)

    if lowered.endswith(".pdf"):
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(raw))
        return "\n".join(page.extract_text() or "" for page in reader.pages)

    return None


async def _save_and_reply(message: Message, profile: dict, report: llm.FoodReport,
                          source: str, raw_input: str | None,
                          guide_url: str = "") -> None:
    """Общий хвост для всех форматов: пишем в БД и показываем итог дня."""
    if not report.usable:
        await message.answer(
            report.comment
            or "Не смог разобрать, что это. Напиши калории текстом — например, «1800 ккал»."
        )
        return

    log_date = today_msk()
    protein_g, fat_g, carb_g = report.macros
    await db.add_entry(
        tg_id=message.from_user.id,
        log_date=log_date,
        kcal=report.kcal,
        protein_g=protein_g,
        fat_g=fat_g,
        carb_g=carb_g,
        is_full_day=report.is_full_day,
        source=source,
        raw_input=(raw_input or "")[:4000] or None,
        llm_note=(report.note or report.comment) or None,
    )

    totals = await db.day_totals(message.from_user.id, log_date)
    # Про оценку по фото человек через секунду прочитает отдельное и куда более
    # внятное предупреждение — вежливое «это прикидка» в подписи только дублирует.
    photo_estimate = report.estimated_from_photo and source == "photo"

    # note объясняет, откуда взялась цифра, — это важнее вежливого comment,
    # потому что именно по нему человек заметит, что данные прочитаны не так.
    footnote = report.note or ("" if photo_estimate else report.comment)
    if report.note:
        # Цифру не прочли напрямую, а вывели — покажем путь к исправлению.
        footnote += "\nЕсли это не так — /reset и пришли данные заново."
    await message.answer(_day_report(profile, totals, footnote))

    # Отдельным сообщением, а не припиской к итогу: предупреждение длинное и в
    # хвосте отчёта его дочитывают через раз.
    if photo_estimate:
        await message.answer(
            _photo_estimate_warning(guide_url), disable_web_page_preview=True
        )


def _day_report(profile: dict, totals, comment: str | None = None) -> str:
    """Итог дня с оценкой попадания в норму."""
    norm = profile["kcal_norm"]
    kcal = float(totals["kcal"])
    diff = kcal - norm
    ratio = kcal / norm if norm else 0

    if 0.9 <= ratio <= 1.1:
        verdict = "✅ В норме — день засчитан."
    elif ratio < 0.9:
        verdict = f"🔻 Недобор {abs(diff):.0f} ккал. Слишком большой дефицит тоже вредит."
    else:
        verdict = f"🔺 Перебор {diff:.0f} ккал."

    lines = [
        f"<b>Сегодня: {kcal:.0f} / {norm} ккал</b>",
        verdict,
        "",
        f"🥩 Белки: {float(totals['protein_g']):.0f} / {profile['protein_g']} г",
        f"🥑 Жиры: {float(totals['fat_g']):.0f} / {profile['fat_g']} г",
        f"🍞 Углеводы: {float(totals['carb_g']):.0f} / {profile['carb_g']} г",
    ]
    if comment:
        lines += ["", f"<i>{comment}</i>"]
    return "\n".join(lines)


@router.message(Command("today"))
async def cmd_today(message: Message) -> None:
    profile = await _require_profile(message)
    if not profile:
        return

    totals = await db.day_totals(message.from_user.id, today_msk())
    if not totals["entries"]:
        await message.answer(
            f"За сегодня записей пока нет.\n"
            f"Твоя норма — {profile['kcal_norm']} ккал. Присылай, что ел."
        )
        return
    await message.answer(_day_report(profile, totals))


@router.message(Command("reset"))
async def cmd_reset(message: Message) -> None:
    profile = await _require_profile(message)
    if not profile:
        return

    deleted = await db.clear_day(message.from_user.id, today_msk())
    if deleted:
        await message.answer(f"Стёр записи за сегодня ({deleted} шт.). Начинаем день заново.")
    else:
        await message.answer("За сегодня и так ничего не записано.")


@router.message(F.document)
async def on_document(message: Message, bot: Bot) -> None:
    profile = await _require_profile(message)
    if not profile:
        return

    document = message.document
    if document.file_size and document.file_size > MAX_FILE_BYTES:
        await message.answer("Файл слишком большой. Пришли выгрузку за неделю или за день.")
        return

    await bot.send_chat_action(message.chat.id, "typing")
    buffer = await bot.download(document.file_id)
    raw = buffer.read()

    try:
        text = _extract_text(document.file_name or "file", raw)
    except Exception:
        log.exception("Не удалось разобрать файл %s", document.file_name)
        await message.answer(
            "Файл не открылся. Попробуй выгрузить в CSV или просто напиши калории текстом."
        )
        return

    if not text or not text.strip():
        await message.answer(
            "Не понимаю этот формат. Подойдут CSV, XLSX, PDF, TXT — "
            "или скриншот дневника картинкой."
        )
        return

    try:
        report = await llm.parse_document(document.file_name or "выгрузка", text)
    except Exception:
        log.exception("LLM не разобрал документ")
        await message.answer("Не смог обработать файл, попробуй ещё раз через минуту.")
        return

    await _save_and_reply(message, profile, report, "document", document.file_name)


@router.message(F.photo)
async def on_photo(message: Message, bot: Bot, cfg: Config) -> None:
    profile = await _require_profile(message)
    if not profile:
        return

    await bot.send_chat_action(message.chat.id, "typing")
    # Берём самое большое превью — мелкий текст в дневнике иначе не читается.
    buffer = await bot.download(message.photo[-1].file_id)

    try:
        report = await llm.parse_image(buffer.read())
    except Exception:
        log.exception("LLM не разобрал изображение")
        await message.answer("Не смог разобрать картинку, попробуй ещё раз или напиши текстом.")
        return

    await _save_and_reply(
        message, profile, report, "photo", message.caption, cfg.guide_url
    )


@router.message(F.text & ~F.text.startswith("/"))
async def on_text(message: Message, bot: Bot) -> None:
    profile = await _require_profile(message)
    if not profile:
        return

    await bot.send_chat_action(message.chat.id, "typing")
    try:
        report = await llm.parse_text(message.text)
    except Exception:
        log.exception("LLM не разобрал текст")
        await message.answer("Не смог обработать сообщение, попробуй ещё раз через минуту.")
        return

    await _save_and_reply(message, profile, report, "text", message.text)
